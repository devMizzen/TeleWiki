import discord
import random
import openpyxl

from discord.ext import commands
from discord.utils import get


TOKEN = 'Njk2NjExNDU5MjIzOTEyNTI4.XorQNA._MZFVS_xMocr9J238zar4g6aOQg'

bot = commands.Bot(command_prefix='!')

strings = [
    "Everybody hide their telescopes!",
    "And has got a HUGE telescope :flushed:",
    "Did you forget to remove dust cap?",
    "Get a goldline!",
    "And is sirius!",
    "Hope you don't have a powerseeker...",
    'Grab an 8" dob with some goldlines and have a seat.',
    'Use me to help you!'

]


@bot.command()
async def commands(ctx):
    embed = discord.Embed(title="TeleWiki Help", description="I know it has been a hard day... I'm here to make it easier :)", color=0x00ff00) 
    embed.add_field(name="**__Exit Pupil:__**", value="Format:\n`!exitpupil <eyepiece fl[mm]> <f-ratio of telescope>`", inline=False)
    embed.add_field(name="**__Magnification:__**",value="Format:\n`!mag <Telescope fl[mm]> <Eyepiece fl[mm]>`",inline=False)
    embed.add_field(name="**__True FOV:__**\n",value="Format:\n`!TFOV <Apparent FOV of eyepiece> <Magnification with eyepiece>`", inline=False)
    embed.add_field(name="**__Barlow:__**\n", value="Format:\n`!barlow <Barlow mag> <fl of ep1> <fl of ep2>...<fl of ep'n'>`\nExample: `!barlow 2 6 9 25`", inline=False)
    embed.add_field(name="**__Gear Database:__**\n", value="Format:\n`!gear help`",inline=False)
    
    await ctx.send(embed=embed)

@bot.command()
async def exitpupil(ctx, *args):
	if args[0] == 'help' or len(args) not in range(1,3):
		embed = discord.Embed(title = "`!exitpupil`")
		embed.add_field(name="Format:", value="`!exitpupil <eyepiece fl[mm]> <f-ratio of telescope>`", inline=False)
		await ctx.send(embed=embed)
		return
		
	try:
		afov = float(args[0])
		mag = float(args[1])
	except:
		embed = discord.Embed(title = "`!exitpupil`")
		embed.add_field(name="Format:", value="`!exitpupil <eyepiece fl[mm]> <f-ratio of telescope>`", inline=False)		
		await ctx.send(embed = embed)
		return
		
	if afov/mag > 7 or afov/mag <0.5:
		col = 0xff0000
	elif afov/mag == 7 or (afov/mag >= 0.5 and afov/mag <=1):
		col = 0xffff00
	else:
		col = 0x00ff00
		
	arg1 = afov
	arg2 = mag
	
	embed = discord.Embed(title="Exit Pupil", description="You would want to keep it below 7mm as average human eye can only expand upto 7mm (decreases as age increases)", color=col)
	embed.add_field(name="Eyepiece focal length", value=str(arg1)+'mm', inline=False)
	embed.add_field(name="Telescope f-ratio", value='F/'+str(arg2), inline=False)
	embed.add_field(name="**`Exit Pupil`**", value=str(round(arg1/arg2,2))+'mm', inline=False)
	await ctx.send(embed=embed)

		
'''@bot.command()
async def exitpupil(ctx, arg1: float, arg2: float):
    if arg1/arg2 > 7 or arg1/arg2 <0.5:
        col = 0xff0000
    elif arg1/arg2 == 7 or (arg1/arg2 >= 0.5 and arg1/arg2 <=1):
        col = 0xffff00
    else:
        col = 0x00ff00
    embed = discord.Embed(title="Exit Pupil", description="You would want to keep it below 7mm as average human eye can only expand upto 7mm (decreases as age increases)", color=col)
    embed.add_field(name="Eyepiece focal length", value=str(arg1)+'mm', inline=False)
    embed.add_field(name="Telescope f-ratio", value='F/'+str(arg2), inline=False)
    embed.add_field(name="**`Exit Pupil`**", value=str(round(arg1/arg2,2))+'mm', inline=False)
    await ctx.send(embed=embed)
'''

@bot.command()
async def mag(ctx, *args):
	if args[0] == 'help' or len(args) not in range(1,3):
		embed = discord.Embed(title = "`!mag`")
		embed.add_field(name="Format:", value="`!mag <Telescope fl[mm]> <Eyepiece fl[mm]>`", inline=False)
		await ctx.send(embed=embed)
		return
		
	try:
		afov = float(args[0])
		mag = float(args[1])
	except:
		embed = discord.Embed(title = "`!mag`")
		embed.add_field(name="Format:", value="`!mag <Telescope fl[mm]> <Eyepiece fl[mm]>`", inline=False)
		await ctx.send(embed=embed)
		return
	
	arg1,arg2 = afov,mag
	
	if arg1/arg2 > 200:
		col = 0xffff00
	else:
		col = 0x00ff00
	embed = discord.Embed(title="Magnification", description="Power of your telescope with a specific eyepiece", color = col)
	embed.add_field(name="Telescope Focal length", value=str(arg1)+'mm', inline=False)
	embed.add_field(name="Eyepiece focal length", value=str(arg2)+'mm', inline=False)
	embed.add_field(name="**`Magnification`**", value=(str(round(arg1/arg2,2))+'X'), inline=False)
	await ctx.send(embed=embed)


@bot.command()
async def TFOV(ctx, *args):
	if args[0] == 'help' or len(args) not in range(1,3):
		embed = discord.Embed(title = "`!True FOV`")
		embed.add_field(name="Format:", value="`!TFOV <Eyepiece fl[mm]> <f-ratio of telescope>`", inline=False)
		await ctx.send(embed=embed)
		return
		
	try:
		afov = float(args[0])
		mag = float(args[1])
	except:
		embed = discord.Embed(title = "`!mag`")
		embed.add_field(name="Format:", value="`!mag <Telescope fl[mm]> <Eyepiece fl[mm]>`", inline=False)
		await ctx.send(embed=embed)
		return
	
	arg1,arg2 = afov,mag
	
	embed = discord.Embed(title="True FOV", description="The amount of sky that you see through the eyepiece", color=0x00ff00)
	embed.add_field(name="Apparent FOV", value=str(arg1)+' degrees', inline=False)
	embed.add_field(name="Magnification", value=str(arg2)+'X', inline=False)
	embed.add_field(name="**`True FOV`**", value=str(round(arg1/arg2,2)), inline=False)
	embed.add_field(name="*Number of full moon*",value=(round((arg1/arg2)/0.5,2)),inline=False)
	await ctx.send(embed=embed)

@bot.command()
async def barlow(ctx, *args):
	if args[0] == 'help' or len(args) == 0:
		embed = discord.Embed(title = "`!barlow`")
		embed.add_field(name="Format:", value="`!barlow <barlow power>  <fl of ep1> <fl of ep2>...<fl of ep'n'>`\nExample: `!barlow 2 6 9 25`", inline=False)
		await ctx.send(embed=embed)
		return
		
	try:
		for arg in range(len(args)):
			args[arg] = float(args[arg])
	except:
		embed = discord.Embed(title = "`!barlow`")
		embed.add_field(name="Format:", value="`!barlow <barlow power>  <fl of ep1> <fl of ep2>...<fl of ep'n'>`\nExample: `!barlow 2 6 9 25`", inline=False)
		await ctx.send(embed=embed)
		return

	barlowedEP = []
	count = 0
	for eyepiece in args:
		if count == 0:
			count += 1
			continue
		else:
			barlowedEP.append(eyepiece)
	
	embed = discord.Embed(title="Barlow Collection", decription="This function takes your barlow and eyepieces and tells the possibe focal lengths.", color = 0x00ff00)
	for lengths in barlowedEP:
		embed.add_field(name='** '+str(lengths)+"mm**", value = str(args[0])+"x *barlowed*: "+"**"+str(round(lengths/args[0],2))+'mm**', inline = True)
	await ctx.send(embed=embed)

@bot.command()
async def gear(ctx, *args):
	if args[0] == 'help':
		embed = discord.Embed(title = "Gear Database Format", description="This database can store information about your rig.\n\
            Please follow the following format:",color=0x00ff00)
		embed.add_field(name="**Telescope**",value="\
            **__New:__**\n\t\
            `   !gear scope new <Aperture[in mm]> <Focal Length[in mm]>`\n\
            **__List:__**\n\t\
            `   !gear scope list`\n\
            **__Remove:__**\n\t\
            `   !gear scope remove <Telescope ID(See in list)>`",inline=False)
		embed.add_field(name="**Eyepiece**",value="\
            **__New:__**\n\t\
            `   !gear ep new <Focal Length[in mm]> <Apparent FOV>`\n\
            **__List:__**\n\t\
            `   !gear ep list`\n\
            **__Remove:__**\n\t\
            `   !gear ep remove <Eyepiece ID(See in list)>`",inline=False)
		embed.add_field(name="**Barlow**",value="\
            **__New:__**\n\t\
            `   !gear barlow new <Power>`\n\
            **__List:__**\n\t\
            `   !gear barlow list`\n\
            **__Remove:__**\n\t\
            `   !gear barlow remove <Barlow ID(See in list)>`",inline=False)
		embed.add_field(name="**Rig**", value="\
			To view the whole rig together: `!gear list`", inline=False)
		await ctx.send(embed=embed)
	elif args[0] == 'scope':
		if args[1] == 'new':
			try:
				int(args[3])/int(args[2])
			except:
				print(ctx.author.name,"entered telescope wrongly")
				msg = "Please only enter unitless values of Aperture and Focal Length"
				await ctx.send(msg)
				return
			
			try:
				wb = openpyxl.load_workbook('Telescopes.xlsx')
				ws = wb.active
				
				aperture = float(args[2])
				focalLength = float(args[3])
				fRatio = focalLength/aperture
				
				f = open("count_telescope.txt",'r')
				count = int(f.read())
				f.close()
				
				ws.append([str(ctx.author.id), ctx.author.name, aperture, focalLength, focalLength/aperture, str(count)])
				wb.save("Telescopes.xlsx")
				
				newCount = count + 1
				f = open("count_telescope.txt",'w')
				f.write(str(newCount))
				f.close()
				
				msg = discord.Embed(title=":white_check_mark:", description = "New telescope added succesfully!", color=0x00ff00)
				await ctx.send(embed=msg)
			except:
				msg = discord.Embed(title=":x:", description="An error occured, Please contact <@362759603768918018>", color=0xff0000)
				await ctx.send(embed=msg)
		
		elif args[1] == 'list':
			wb = openpyxl.load_workbook('Telescopes.xlsx')
			ws = wb.active
			f=open('count_telescope.txt','r')
			count = int(f.read())
			entries = []
			
			for x in range(1,count+1):
                #print(ws.cell(row=x,column=1).value)
                #print(ctx.author.id)
				entries.append(x)
				
			ctr = 0
			embed = discord.Embed(title="Your Scopes",description= "If you have added any scope, it will appear here:", color=0x00ff00, inline = False)
			for n in range(len(entries)):
				if str(ws.cell(row=n+1,column=1).value) == str(ctx.author.id):
					embed.add_field(name=("**Telescope "+str(ctr+1)+"**"), value=("**__Aperture:__**"+" "+str(ws["C"+str(entries[n])].value)+"mm\n**__Focal Length:__**"+" "+str(ws["D"+str(entries[n])].value)+"mm\n*Telescope ID:* "+str(ws["F"+str(entries[n])].value)))
					ctr += 1
			await ctx.send(embed=embed)
			
		elif args[1] == 'remove':
			wb = openpyxl.load_workbook("Telescopes.xlsx")
			ws = wb.active
			
			f = open("count_telescope.txt","r")
			count = int(f.read())
			f.close()
			for n in range(count):
				if str(args[2]) == str(ws.cell(row=n+1,column=6).value):
					if str(ws.cell(row=n+1,column=1).value) == str(ctx.author.id):
                        #movestring = "A"+str(n+1)+":F"+str(count)
                        #ws.move_range(movestring, rows=-1, cols=0)
						for a in range(n, count):
							ws["A"+str(a+1)] = ws.cell(row=a+2,column=1).value
							ws["B"+str(a+1)] = ws.cell(row=a+2,column=2).value
							ws["C"+str(a+1)] = ws.cell(row=a+2,column=3).value
							ws["D"+str(a+1)] = ws.cell(row=a+2,column=4).value
							ws["E"+str(a+1)] = ws.cell(row=a+2,column=5).value
							ws["F"+str(a+1)] = ws.cell(row=a+2,column=6).value
							
						wb.save("Telescopes.xlsx")
						newCount = count - 1
						f = open("count_telescope.txt","w")
						f.write(str(newCount))
						f.close()
						embed = discord.Embed(name=':white_check_mark:', description="Telescope deleted succesfully!", color=0x00ff00)
						await ctx.send(embed=embed)
						error = False
					else:
						embed = discord.Embed(name=':x:', description="You cannot delete someone else's scope!",color=0xff0000)
						await ctx.send(embed=embed)
						error = False
				else:
					error = True
					continue
				if error:
					await ctx.send("Wrong ID")
				else:
					return
	
	elif args[0] == 'ep':
		if args[1] == 'new':
			try:
				float(args[3])/float(args[2])
			except:
				print(ctx.author.name,"entered eyepiece wrongly")
				msg = "Please only enter unitless values of Focal Length and Apparent FOV"
				await ctx.send(msg)
				return
			
			try:
				wb = openpyxl.load_workbook('Eyepieces.xlsx')
				ws = wb.active
				
				focalLength = float(args[2])
				afov = float(args[3])
				
				try:
					f = open("count_eyepiece.txt",'r')
					count = int(f.read())
					f.close()
				except:
					count = 0
#                                A                   B               C          D        E
				ws.append([str(ctx.author.id), ctx.author.name, focalLength, afov, str(count)])
				wb.save("Eyepieces.xlsx")
				
				newCount = count + 1
				f = open("count_eyepiece.txt",'w')
				f.write(str(newCount))
				f.close()
				
				msg = discord.Embed(title=":white_check_mark:", description = "New eyepiece added succesfully!", color=0x00ff00)
				await ctx.send(embed=msg)
			except:
				msg = discord.Embed(title=":x:", description="An error occured, Please contact <@362759603768918018>", color=0xff0000)
				await ctx.send(embed=msg)
		elif args[1] == 'list':
			wb = openpyxl.load_workbook('Eyepieces.xlsx')
			ws = wb.active
			f=open('count_eyepiece.txt','r')
			count = int(f.read())
			entries = []
			
			for x in range(1,count+1):
                #print(ws.cell(row=x,column=1).value)
                #print(ctx.author.id)
				entries.append(x)
			
			ctr = 0
			embed = discord.Embed(title="Your Eyepiece",description= "If you have added any eyepiece, it will appear here:", color=0x00ff00, inline = False)
			for n in range(len(entries)):
				if str(ws.cell(row=n+1,column=1).value) == str(ctx.author.id):
					embed.add_field(name=("**Eyepiece "+str(ctr+1)+"**"), value=("**__Focal Lenght:__**"+" "+str(ws["C"+str(entries[n])].value)+"mm\n**__Apparent FOV:__**"+" "+str(ws["D"+str(entries[n])].value)+"°\n*Eyepiece ID:* "+str(ws["E"+str(entries[n])].value)))
					ctr += 1
			await ctx.send(embed=embed)
		
		elif args[1] == 'remove':
			wb = openpyxl.load_workbook("Eyepieces.xlsx")
			ws = wb.active
			f = open("count_eyepiece.txt","r")
			count = int(f.read())
			f.close()
			for n in range(count):
				if str(args[2]) == str(ws.cell(row=n+1,column=5).value):
					if str(ws.cell(row=n+1,column=1).value) == str(ctx.author.id):
                        #movestring = "A"+str(n+1)+":F"+str(count)
                        #ws.move_range(movestring, rows=-1, cols=0)
						for a in range(n, count):
							ws["A"+str(a+1)] = ws.cell(row=a+2,column=1).value
							ws["B"+str(a+1)] = ws.cell(row=a+2,column=2).value
							ws["C"+str(a+1)] = ws.cell(row=a+2,column=3).value
							ws["D"+str(a+1)] = ws.cell(row=a+2,column=4).value
							ws["E"+str(a+1)] = ws.cell(row=a+2,column=5).value
						wb.save("Eyepieces.xlsx")
						newCount = count - 1
						f = open("count_eyepiece.txt","w")
						f.write(str(newCount))
						f.close()
						embed = discord.Embed(name=':white_check_mark:', description="Eyepiece deleted succesfully!", color=0x00ff00)
						await ctx.send(embed=embed)
						error = False
					else:
						embed = discord.Embed(name=':x:', description="You cannot delete someone else's eyepiece!",color=0xff0000)
						await ctx.send(embed=embed)
						error = False
				else:
					error = True
					continue
				if error:
					await ctx.send("Wrong ID")
				else:
					return
					
	elif args[0] == 'barlow':
		if args[1] == 'new':
			try:
				testVar = float(args[2])
			except:
				print(ctx.author.name,"entered barlow wrongly")
				msg = "Please only enter unitless value of barlow."
				await ctx.send(msg)
				return

			try:
				wb = openpyxl.load_workbook('Barlows.xlsx')
				ws = wb.active
				
				power = args[2]
				
				try:
					f = open("count_barlow.txt",'r')
					count = int(f.read())
					f.close()
				except:
					count = 0
#                                A                   B            C         D
				ws.append([str(ctx.author.id), ctx.author.name, power, str(count)])
				wb.save("barlows.xlsx")
				
				newCount = count + 1
				f = open("count_barlow.txt",'w')
				f.write(str(newCount))
				f.close()
				
				msg = discord.Embed(title=":white_check_mark:", description = "New barlow added succesfully!", color=0x00ff00)
				await ctx.send(embed=msg)
			except:
				msg = discord.Embed(title=":x:", description="An error occured, Please contact <@362759603768918018>", color=0xff0000)
				await ctx.send(embed=msg)
		elif args[1] == 'list':
			wb = openpyxl.load_workbook('Barlows.xlsx')
			ws = wb.active
			f=open('count_barlow.txt','r')
			count = int(f.read())
			entries = []
			
			for x in range(1,count+1):
                #print(ws.cell(row=x,column=1).value)
                #print(ctx.author.id)
				entries.append(x)
            
			ctr = 0
			embed = discord.Embed(title="Your Barlows",description= "If you have added any barlow, it will appear here:", color=0x00ff00, inline = False)
			for n in range(len(entries)):
				if str(ws.cell(row=n+1,column=1).value) == str(ctx.author.id):
					embed.add_field(name=("**Barlow "+str(ctr+1)+"**"), value=("**__Power:__**"+" "+str(ws["C"+str(entries[n])].value)+"X\n*Barlow ID:* "+str(ws["D"+str(entries[n])].value)))
					ctr += 1
			await ctx.send(embed=embed)

		elif args[1] == 'remove':
			wb = openpyxl.load_workbook("Barlows.xlsx")
			ws = wb.active
			f = open("count_barlow.txt","r")
			count = int(f.read())
			f.close()
			for n in range(count):
				if str(args[2]) == str(ws.cell(row=n+1,column=4).value):
					if str(ws.cell(row=n+1,column=1).value) == str(ctx.author.id):
                        #movestring = "A"+str(n+1)+":F"+str(count)
                        #ws.move_range(movestring, rows=-1, cols=0)
						for a in range(n, count):
							ws["A"+str(a+1)] = ws.cell(row=a+2,column=1).value
							ws["B"+str(a+1)] = ws.cell(row=a+2,column=2).value
							ws["C"+str(a+1)] = ws.cell(row=a+2,column=3).value
							ws["D"+str(a+1)] = ws.cell(row=a+2,column=4).value
						
						wb.save("Barlows.xlsx")
						newCount = count - 1
						f = open("count_barlow.txt","w")
						f.write(str(newCount))
						f.close()
						embed = discord.Embed(name=':white_check_mark:', description="Barlow deleted succesfully!", color=0x00ff00)
						await ctx.send(embed=embed)
						error = False
					else:
						embed = discord.Embed(name=':x:', description="You cannot delete someone else's barlow!",color=0xff0000)
						await ctx.send(embed=embed)
						error = False
				else:
					error = True
					continue
				if error:
					await ctx.send("Wrong ID")
				else:
					return
				
	elif args[0] == 'list':

		if len(args) >= 2:
			user = ""
			for arg in range(len(args)-1):
				user += args[arg+1]
				user += ' '
				
			user = user.lower()
			for member in ctx.guild.members:
				if user == str(member.name+' ').lower() or user == str(member.nick+' ').lower():
					user = member
					break
			
			try:
				chk = member.name
			except:
				msg = user + "is not a valid user"
				await ctx.send(msg)

		else:
			member = ctx.author
		'''try:
			user = ''
			for arg in range(len(args)+1):
				user += args[arg+1]
			for member in ctx.guild.members:
				if user == member.name or user == member.nick:
					user = member
					break
		except:
			user = ctx.author'''

		Title = (str(member.name) + "'s Rig")
		
		f = open("count_telescope.txt", 'r')
		Tcount = int(f.read())
		f.close()
		
		f = open("count_eyepiece.txt", 'r')
		Ecount = int(f.read())
		f.close()
		
		f = open("count_barlow.txt", 'r')
		Bcount = int(f.read())
		f.close()
		
		ctr = 0
		ctr_ep = 0
		ctr_br = 0

		ctr_title = 0
		
		for telescope in range(Tcount):
			wb = openpyxl.load_workbook("Telescopes.xlsx")
			ws = wb.active
			
			if str(member.id) == ws.cell(row=telescope+1, column=1).value:
				ap = round(float(ws.cell(row=telescope+1, column=3).value), 2)
				fl_scope = round(float(ws.cell(row=telescope+1, column=4).value), 2)
				ratio = round(float(ws.cell(row=telescope+1, column=5).value), 2)
				
				telescope_emb = discord.Embed(title = "**__Telescope: "+str(ctr+1)+"__**", description = ("\n**Aperture** = "+str(ap)+"mm"+"\n\
**Focal Length** = "+str(fl_scope)+"mm\n\
**f-ratio**: "+str(ratio)), color = 0x00FF00)
				ctr += 1
				
				for eyepiece in range(Ecount):	
					wb = openpyxl.load_workbook("Eyepieces.xlsx")
					ws = wb.active
					
					if str(member.id) == ws.cell(row=eyepiece+1, column=1).value:
						fl_ep = float(ws.cell(row=eyepiece+1,column=3).value)
						afov = float(ws.cell(row=eyepiece+1,column=4).value)
						
						mag = round(fl_scope/fl_ep, 2)
						tfov = round(afov/mag, 2)
						moons = round(tfov/0.5, 2)
						ep = round(fl_ep/ratio, 2)
						
						telescope_emb.add_field(name=("**	__Eyepiece "+str(ctr_ep+1)+":__** `"+str(fl_ep)+"mm`"), value = ("\n**		Magnification** = "+str(mag)+"X\n\
**		True FOV** = "+str(tfov)+"° ; *No. of moons = "+str(moons)+"*\n\
**		Exit Pupil** = "+str(ep)+"mm"), inline = False)
						ctr_ep += 1
						
						for barlow in range(Bcount):
							wb = openpyxl.load_workbook("Barlows.xlsx")
							ws = wb.active
							
							if str(member.id) == ws.cell(row=barlow+1, column=1).value:
								power = float(ws.cell(row=barlow+1, column=3).value)
								fl_scopenew = round(fl_scope*power, 2)
								
								mag = round(fl_scopenew/fl_ep, 2)
								tfov = round(afov/mag, 2)
								moons = round(tfov/0.5, 2)
								ep = round((fl_ep/power)/ratio, 2)
								
								telescope_emb.add_field(name=("**	Eyepiece "+str(ctr_ep)+"** *(barlowed):* `"+str(round(fl_ep/power, 2))+"mm`"), value = ("\n\
**		Magnification** = "+str(mag)+"X\n\
**		True FOV** = "+str(tfov)+"° ; *No. of moons = "+str(moons)+"*\n\
**		Exit Pupil** = "+str(ep)+"mm"), inline = False)
								
								ctr_br += 1
							else:
								continue
					
					else:
						continue
				if ctr_title == 0:
					await ctx.send(Title)
					ctr_title += 1
				await ctx.send(embed = telescope_emb)
								
			else:
				continue

		if ctr==0:
			if user == ctx.author:
				await ctx.send("You don't have any gear!\nDo `!gear help` to know more")
			else:
				msg = member.name + " Doesn't have any gear!"
				await ctx.send(msg)
			
@bot.event
async def on_message(ctx):
	if ctx.author == bot.user:
		return
	elif 'powerseeker' in ctx.content.lower() or 'power seeker' in ctx.content.lower():
		link = "https://www.reddit.com/r/dontbuyapowerseeker/comments/7m8d8k/why_each_powerseeker_sucks/?utm_medium=android_app&utm_source=share"
		await ctx.channel.send("OMG its the p-word.\n"+link)
	else:
		await bot.process_commands(ctx)

@bot.event
async def on_member_join(member):
    rstring = strings[random.randint(0,len(strings))]
    regionchannel = bot.get_channel(543251534104756254)
    for channel in member.guild.channels:
    	if str(channel) == "general":
    		await channel.send(f"{member.mention} has joined the server! "+rstring+f"\nHead to the {regionchannel.mention} to set your region.")

@bot.command()
async def region(ctx, *reg):
    try:
        a = reg[1]
        entered_role = reg[0]+' '+reg[1]
    except:
        entered_role = reg[0]
    role = discord.utils.get(ctx.guild.roles, name=entered_role)
    roles = [
	'542615298742943755', #Australia
    '543250771358121994', #Europe
    '543250774314975242', #United States
    '543250778236649483', #Zealand you freedum burger
    '543250776534024193', #Central America
    '543250777465159680', #South Amarica
    '550717762062778370', #Africa
    '696312571698872371', #Asia
    '543250775795695616' #Canada
	]
    for r in ctx.author.roles:
        if r.position>=1:
            await ctx.channel.send("You already have a region assigned. If you want to change region, please contact a moderator.")
            return
    if entered_role == 'New Zealand':
        await ctx.author.add_roles(roles[3])
    if role is None:
        await ctx.channel.send(
        "Region doesn't exist. List of available regions:\
        \nAustralia\
		\nEurope\
		\nUnited States\
		\nNew Zealand\
		\nCentral America\
		\nSouth America\
		\nAfrica\
		\nAsia\
		\nCanada\n\
		\nIf your region doesn't exist in the list, please contact a moderator."
		)
    elif role in ctx.author.roles:
        await ctx.channel.send("You already have this region assigned to you.")
    else:
        try:
            await ctx.author.add_roles(role)
            await ctx.channel.send("Region assigned succesfully!")
        except:
            await ctx.channel.send("Some error occured, please contact <@362759603768918018>.")

@bot.event
async def on_ready():
    print('Logged in as')
    print(bot.user.name)
    print(bot.user.id)
    print('------')

bot.run(TOKEN)
